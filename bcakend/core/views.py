from rest_framework import viewsets, status, generics
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from rest_framework.authtoken.models import Token
from django.contrib.auth import authenticate
from django.contrib.auth.models import User
from django.db.models import Count, Q
from django.utils import timezone
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date, timedelta

from .models import Participant, Phlebotomy, SampleProcessing, SampleStorage, StockItem, AuditLog
from .serializers import (
    ParticipantSerializer, PhlebotomySerializer, SampleProcessingSerializer,
    SampleStorageSerializer, StockItemSerializer, AuditLogSerializer, UserSerializer
)


def log_action(user, action, instance):
    AuditLog.objects.create(
        user=user, action=action,
        model_name=instance.__class__.__name__,
        object_id=str(instance.pk),
        object_repr=str(instance),
    )


# ── AUTH ─────────────────────────────────────
@api_view(['POST'])
@permission_classes([AllowAny])
def login_view(request):
    username = request.data.get('username')
    password = request.data.get('password')
    user = authenticate(username=username, password=password)
    if user:
        token, _ = Token.objects.get_or_create(user=user)
        return Response({'token': token.key, 'user': UserSerializer(user).data})
    return Response({'error': 'Invalid credentials'}, status=status.HTTP_401_UNAUTHORIZED)


@api_view(['POST'])
def logout_view(request):
    request.user.auth_token.delete()
    return Response({'message': 'Logged out'})


@api_view(['GET'])
def me_view(request):
    return Response(UserSerializer(request.user).data)


# ── DASHBOARD ────────────────────────────────
@api_view(['GET'])
def dashboard_stats(request):
    today = date.today()
    last_30 = today - timedelta(days=30)

    # Sample type distribution
    sample_dist = (
        Phlebotomy.objects
        .values('sample_type')
        .annotate(count=Count('id'))
    )

    # Collections per day (last 14 days)
    daily = []
    for i in range(13, -1, -1):
        d = today - timedelta(days=i)
        cnt = Phlebotomy.objects.filter(collection_date=d).count()
        daily.append({'date': str(d), 'count': cnt})

    # Stock expiry warnings
    expiring_soon = StockItem.objects.filter(
        expiry_date__lte=today + timedelta(days=30),
        expiry_date__gte=today
    ).count()

    expired = StockItem.objects.filter(expiry_date__lt=today).count()

    stats = {
        'total_participants': Participant.objects.count(),
        'total_samples': Phlebotomy.objects.count(),
        'total_processings': SampleProcessing.objects.count(),
        'total_storage': SampleStorage.objects.count(),
        'total_stock_items': StockItem.objects.count(),
        'samples_this_month': Phlebotomy.objects.filter(collection_date__gte=last_30).count(),
        'participants_this_month': Participant.objects.filter(enrollment_date__gte=last_30).count(),
        'stock_expiring_soon': expiring_soon,
        'stock_expired': expired,
        'sample_type_distribution': list(sample_dist),
        'daily_collections': daily,
    }
    return Response(stats)


# ── EXPORT HELPERS ───────────────────────────
def make_excel_response(filename):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def style_header(ws, headers):
    header_fill = PatternFill('solid', start_color='1A6B8A', end_color='1A6B8A')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')


# ── VIEWSETS ─────────────────────────────────
class ParticipantViewSet(viewsets.ModelViewSet):
    queryset = Participant.objects.all()
    serializer_class = ParticipantSerializer
    search_fields = ['participant_id', 'study_name']
    filterset_fields = ['sex', 'study_name']
    ordering_fields = ['enrollment_date', 'age', 'participant_id']

    def perform_create(self, serializer):
        instance = serializer.save(created_by=self.request.user)
        log_action(self.request.user, 'create', instance)

    def perform_update(self, serializer):
        instance = serializer.save()
        log_action(self.request.user, 'update', instance)

    def perform_destroy(self, instance):
        log_action(self.request.user, 'delete', instance)
        instance.delete()

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Participants'
        headers = ['Participant ID', 'Study Name', 'Date of Birth', 'Age', 'Sex', 'Enrollment Date']
        style_header(ws, headers)
        for p in Participant.objects.all():
            ws.append([p.participant_id, p.study_name, str(p.date_of_birth), p.age, p.get_sex_display(), str(p.enrollment_date)])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 20
        response = make_excel_response('participants.xlsx')
        wb.save(response)
        return response


class PhlebotomyViewSet(viewsets.ModelViewSet):
    queryset = Phlebotomy.objects.select_related('participant').all()
    serializer_class = PhlebotomySerializer
    search_fields = ['participant__participant_id', 'collector_name', 'tube_type']
    filterset_fields = ['sample_type', 'tube_type', 'visit_type', 'consented', 'sample_collected']
    ordering_fields = ['collection_date', 'collection_time']

    def perform_create(self, serializer):
        instance = serializer.save(created_by=self.request.user)
        log_action(self.request.user, 'create', instance)

    def perform_update(self, serializer):
        instance = serializer.save()
        log_action(self.request.user, 'update', instance)

    def perform_destroy(self, instance):
        log_action(self.request.user, 'delete', instance)
        instance.delete()

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Phlebotomy'
        headers = ['Participant ID', 'Collector', 'Date', 'Time', 'Sample Type', 'Tube Type', 'Volume', 'Site', 'Notes', 'Consented', 'Visit Type', 'Collected']
        style_header(ws, headers)
        for p in Phlebotomy.objects.select_related('participant').all():
            ws.append([
                p.participant.participant_id, p.collector_name, str(p.collection_date), str(p.collection_time),
                p.get_sample_type_display(), p.get_tube_type_display(), p.get_volume_collected_display(),
                p.get_collection_site_display(), p.get_collection_notes_display(),
                'Yes' if p.consented else 'No', p.get_visit_type_display(), 'Yes' if p.sample_collected else 'No'
            ])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18
        response = make_excel_response('phlebotomy.xlsx')
        wb.save(response)
        return response


class SampleProcessingViewSet(viewsets.ModelViewSet):
    queryset = SampleProcessing.objects.select_related('phlebotomy__participant').all()
    serializer_class = SampleProcessingSerializer
    search_fields = ['accession_number', 'technologist_initials', 'phlebotomy__participant__participant_id']
    filterset_fields = ['processing_type', 'equipment_used']
    ordering_fields = ['reception_date', 'accession_number']

    def perform_create(self, serializer):
        instance = serializer.save(created_by=self.request.user)
        log_action(self.request.user, 'create', instance)

    def perform_update(self, serializer):
        instance = serializer.save()
        log_action(self.request.user, 'update', instance)

    def perform_destroy(self, instance):
        log_action(self.request.user, 'delete', instance)
        instance.delete()

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Sample Processing'
        headers = ['Accession #', 'Participant ID', 'Reception Date', 'Reception Time', 'Processing Type', 'Aliquot #', 'Technologist', 'Equipment', 'Results Dispatched To']
        style_header(ws, headers)
        for sp in SampleProcessing.objects.select_related('phlebotomy__participant').all():
            ws.append([
                sp.accession_number, sp.phlebotomy.participant.participant_id,
                str(sp.reception_date), str(sp.reception_time),
                sp.get_processing_type_display(), sp.aliquot_number,
                sp.technologist_initials, sp.get_equipment_used_display(), sp.results_dispatched_to
            ])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 20
        response = make_excel_response('sample_processing.xlsx')
        wb.save(response)
        return response


class SampleStorageViewSet(viewsets.ModelViewSet):
    queryset = SampleStorage.objects.select_related('processing__phlebotomy__participant').all()
    serializer_class = SampleStorageSerializer
    search_fields = ['sample_id', 'freezer_id', 'fridge_id', 'processing__accession_number']
    filterset_fields = ['storage_temperature', 'storage_condition']
    ordering_fields = ['date_stored', 'sample_id']

    def perform_create(self, serializer):
        instance = serializer.save(created_by=self.request.user)
        log_action(self.request.user, 'create', instance)

    def perform_update(self, serializer):
        instance = serializer.save()
        log_action(self.request.user, 'update', instance)

    def perform_destroy(self, instance):
        log_action(self.request.user, 'delete', instance)
        instance.delete()

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Sample Storage'
        headers = ['Sample ID', 'Accession #', 'Participant ID', 'Freezer ID', 'Shelf', 'Rack', 'Box', 'Position', 'Temperature', 'Date Stored', 'Condition']
        style_header(ws, headers)
        for s in SampleStorage.objects.select_related('processing__phlebotomy__participant').all():
            ws.append([
                s.sample_id, s.processing.accession_number,
                s.processing.phlebotomy.participant.participant_id,
                s.freezer_id, s.shelf_number, s.rack_number, s.box_number, s.position,
                s.get_storage_temperature_display(), str(s.date_stored), s.get_storage_condition_display()
            ])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18
        response = make_excel_response('sample_storage.xlsx')
        wb.save(response)
        return response


class StockItemViewSet(viewsets.ModelViewSet):
    queryset = StockItem.objects.all()
    serializer_class = StockItemSerializer
    search_fields = ['item_id', 'item_name', 'supplier', 'batch_number']
    filterset_fields = ['category', 'storage_location', 'condition_received', 'unit']
    ordering_fields = ['item_name', 'expiry_date', 'quantity_available']

    def perform_create(self, serializer):
        instance = serializer.save(created_by=self.request.user)
        log_action(self.request.user, 'create', instance)

    def perform_update(self, serializer):
        instance = serializer.save()
        log_action(self.request.user, 'update', instance)

    def perform_destroy(self, instance):
        log_action(self.request.user, 'delete', instance)
        instance.delete()

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Stock Inventory'
        headers = ['Item ID', 'Item Name', 'Category', 'Supplier', 'Batch #', 'Expiry Date', 'Qty Available', 'Unit', 'Location', 'Condition', 'Received By', 'Reception Date']
        style_header(ws, headers)
        for item in StockItem.objects.all():
            ws.append([
                item.item_id, item.item_name, item.get_category_display(),
                item.supplier, item.batch_number, str(item.expiry_date),
                float(item.quantity_available), item.get_unit_display(),
                item.get_storage_location_display(), item.get_condition_received_display(),
                item.received_by, str(item.reception_date)
            ])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 20
        response = make_excel_response('stock_inventory.xlsx')
        wb.save(response)
        return response

    @action(detail=False, methods=['get'])
    def low_stock(self, request):
        items = StockItem.objects.filter(quantity_available__lte=10)
        return Response(StockItemSerializer(items, many=True).data)

    @action(detail=False, methods=['get'])
    def expiring(self, request):
        threshold = date.today() + timedelta(days=30)
        items = StockItem.objects.filter(expiry_date__lte=threshold, expiry_date__gte=date.today())
        return Response(StockItemSerializer(items, many=True).data)


class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = AuditLog.objects.select_related('user').all()
    serializer_class = AuditLogSerializer
    filterset_fields = ['action', 'model_name']
    search_fields = ['object_repr', 'user__username']
    ordering_fields = ['timestamp']